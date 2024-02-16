import openpyxl
import os
from openpyxl.styles import NamedStyle
from collections import defaultdict


def list_dict(sheet):
    slov = dict()
    name_firm = dict()
    all_sum = 0
    count = 16
    '''Цикл по составлению словаря из ячеек, где Q это ключ словаря, а Z сумма по объёму сделок контр-агента'''
    while sheet['Z' + str(count)].value is not None :
        if slov.get(sheet['Q' + str(count)].value) is not None:
            value: int = slov.pop(sheet['Q' + str(count)].value)
            slov[sheet['Q' + str(count)].value] = value + int(sheet['Z' + str(count)].value)
        if name_firm.get(sheet['Q' + str(count)].value) is not None:
            if name_firm.get(sheet['Q' + str(count)].value) != [sheet['S' + str(count)].value]:
                name_firm[sheet['Q' + str(count)].value] += [sheet['S' + str(count)].value]
        else:
            slov[sheet['Q' + str(count)].value] = int(sheet['Z' + str(count)].value)
            name_firm[sheet['Q' + str(count)].value] = [sheet['S' + str(count)].value]
        count += 1
    for _,i_value in slov.items():
        all_sum += i_value
    '''Отсортировка компаний с общей суммой контракта за кварталл до 990 000'''
    my_dict = {key: val for key, val in slov.items() if val > 989000 and key != ''}
    dd = defaultdict(list)
    for key in set(my_dict.keys() & name_firm.keys()):
        if key in my_dict:
            dd[key].append(my_dict[key])
        if key in name_firm:
            dd[key].append(name_firm[key])
    return dd, all_sum

accountingStyle = NamedStyle(name='accountingStyle', number_format='#,##0.00')
currencyStyle = NamedStyle(name='currencyStyle', number_format='0')
def func_addfile(name_sheet: str, name, new_dict: dict, all_sum: int):
    '''Функция для создания нового листа в файле и заплнение его значениями словаря'''
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb.create_sheet(name_sheet)
    n = 1
    summa = 0
    for key, val in new_dict.items():
        item = ' '.join(str(el) for el in val[1:])
        if item.strip("[']") == '':
            continue
        else:
            ws['B' + str(n)] = item.strip("[']")
        ws['A' + str(n)] = key
        ws['A' + str(n)].style = currencyStyle
        value = 1000 * round(val[0]/1000)
        summa += value
        ws['C' + str(n)] = value
        ws['C' + str(n)].style = accountingStyle
        n += 1
    ws['D' + str(1)] = summa
    ws['D' + str(1)].style = accountingStyle
    ws['D' + str(2)] = all_sum
    ws['D' + str(2)].style = accountingStyle
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    wb.save('test.xlsx')
    print(f'Файл {name} успешно обработан и записанн в test.xlsx')


files_name = any
path = 'test/'
for root, dirs, files in os.walk(path):
    files_name = files

for i, name in enumerate(files_name):
    kniga = openpyxl.load_workbook(path + files_name[i])
    kniga.active = 0
    sheet = kniga.active

    new_dict, all_sum = list_dict(sheet)
    lst = name.split(' ')
    name_sheet = ''
    if lst[4] != 'продажи.xlsx':
        name_sheet = lst[3] + ' ' + lst[4]
    else:
        name_sheet = lst[3]
    func_addfile(str(name_sheet), name, new_dict, all_sum)
