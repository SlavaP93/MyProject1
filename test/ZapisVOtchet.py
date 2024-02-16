import openpyxl
from openpyxl.styles import NamedStyle
from copy import copy

def style_cell(otchet_firm,new_cell,cell):

    """ Эта функция предназначена для стилизации ячейки new_cell в отчете otchet_firm
    на основе стилей ячейки cell. Если у ячейки cell есть стиль, то стиль new_cell
    копируется из стиля cell (шрифт, заливка,
     выравнивание, защита, границы). Если стиль у cell отсутствует,
    печатается сообщение об ошибке."""

    if otchet_firm[cell].has_style:
        otchet_firm[new_cell].font = copy(otchet_firm[cell].font)
        otchet_firm[new_cell].fill = copy(otchet_firm[cell].fill)
        otchet_firm[new_cell].alignment = copy(otchet_firm[cell].alignment)
        otchet_firm[new_cell].protection = copy(otchet_firm[cell].protection)
        otchet_firm[new_cell].border = copy(otchet_firm[cell].border)
    else:
        print(otchet_firm[cell], 'Стиль клетки не определён',otchet_firm[new_cell])
        return


def func_colpoisk(otchet_firm):

    ''' Эта функция выполняет поиск строки need_row в отчете otchet_firm
    и возвращает букву столбца и позицию предыдущего столбца после нее. Если строка не найдена,
    она создает строку с нужным текстом и стилизирует ячейку с помощью функции style_cell.'''

    need_row = 'Выручка за 4  кв. 2023 г. тыс. руб.    с НДС.'
    need_cell = 'E'
    for row in otchet_firm.iter_rows(min_row=2, max_row=2, min_col=3):
        for cell in row:
            if cell.value == need_row:
                need_letter = cell.column_letter
                left = need_cell.column_letter
                return need_letter, left
            if cell.value is None or cell.value == '':
                cell.value = need_row
                if need_cell.has_style:
                    cell.font = copy(need_cell.font)
                    cell.border = copy(need_cell.border)
                    cell.fill = copy(need_cell.fill)
                    cell.number_format = copy(need_cell.number_format)
                    cell.protection = copy(need_cell.protection)
                    cell.alignment = copy(need_cell.alignment)
                    otchet_firm.column_dimensions[cell.column_letter].width = 18
                need_letter = cell.column_letter
                return need_letter, need_cell.column_letter
            need_cell = cell
    else:
        otchet_firm['K2'] = 'Выручка за 4  кв. 2023 г. тыс. руб.с НДС.'
        style_cell(otchet_firm,'K2','E2')
        otchet_firm.column_dimensions[otchet_firm['K']].width = 18
        return 'K','F'


def func_otchet(firm_name, otchet, test):

    '''Эта функция обрабатывает данные для отчета для конкретной фирмы firm_name.
    Она использует функцию func_colpoisk для поиска необходимой строки,
    затем сравнивает значения из таблиц и еще раз использует функцию style_cell
    для стилизации ячеек в соответствии с условиями. После завершения обработки производит сохранение.'''

    otchet_firm = otchet[firm_name]
    test_firm = test[firm_name]
    colA_test = test_firm['A']
    colC_otchet = otchet_firm['C']
    letter, left_letter = func_colpoisk(otchet_firm)
    otchet_firm.style = currencyStyle
    for ind_test, i_value in enumerate(colA_test):
        for ind_otchet, value_otchet in enumerate(colC_otchet):
            if ind_otchet <= 2:
                continue
            if ind_otchet > 10:
                if value_otchet.value is None:
                    otchet_firm.insert_rows(ind_otchet+1)
                    otchet_firm['B' + str(ind_otchet + 1)] = test_firm['B' + str(ind_test + 1)].value
                    style_cell(otchet_firm,'B' + str(ind_otchet + 1),'B4')
                    otchet_firm['C' + str(ind_otchet + 1)] = test_firm['A' + str(ind_test + 1)].value
                    style_cell(otchet_firm, 'C' + str(ind_otchet + 1), 'C4')
                    otchet_firm[letter + str(ind_otchet + 1)] = test_firm['C' + str(ind_test + 1)].value
                    otchet_firm[letter + str(ind_otchet + 1)].number_format = '#,##0.00'
                    style_cell(otchet_firm, letter + str(ind_otchet + 1), left_letter + str(4))
                    break
            if str(i_value.value) == str(value_otchet.value):
                otchet_firm[letter + str(ind_otchet + 1)] = test_firm['C' + str(ind_test + 1)].value
                style_cell(otchet_firm, letter + str(ind_otchet + 1), left_letter+ str(ind_otchet + 1))
                otchet_firm[letter + str(ind_otchet + 1)].number_format = '#,##0.00'
                otchet_firm.column_dimensions[letter].width = 17
                break
    for ind,row in enumerate(otchet_firm[letter]):
        if ind < 3:
            continue
        if otchet_firm['C' + str(ind + 1)] is None or otchet_firm['C' + str(ind + 1)] == '':
            break
        if row.value is None:
            otchet_firm[letter + str(ind+1)] = 0
            otchet_firm[letter + str(ind + 1)].number_format = '#,##0.00'
            style_cell(otchet_firm, letter + str(ind + 1), left_letter + str(ind + 1))


    for i, row_str in enumerate(otchet_firm['B']):
        if row_str.value == 'Выручка по указанным выше клиентам':
            otchet_firm[letter+str(i+1)] = test_firm['D1'].value
            otchet_firm[letter + str(i + 1)].number_format = '#,##0.00'
            otchet_firm[letter + str(i + 1)].font = openpyxl.styles.Font(bold=True)
            otchet_firm[letter + str(i + 3)] = test_firm['D2'].value
            otchet_firm[letter + str(i + 3)].number_format = '#,##0.00'
            otchet_firm[letter + str(i + 3)].font = openpyxl.styles.Font(bold=True)

    '''Главный цикл: Он загружает отчеты и данные тестов в книги, создает стиль currencyStyle 
        для числовых ячеек, итерирует по именам фирм из файла с данными test.xlsx, 
        обрабатывает их с помощью функции func_otchet, сохраняет результаты в новый отчет и 
        печатает сообщение о завершении обработки. 
        Если имя фирмы не найдено в отчете, записывает ошибку в файл errors.txt.'''

with open('errors.txt', 'w', encoding='UTF-8') as errors_file:
    otchet = openpyxl.load_workbook('2022 2 кв. - 2023г. 2 кв Продажи Общие — копия.xlsx')
    otchet_sheet_names = otchet.sheetnames
    test = openpyxl.load_workbook('test.xlsx')
    test_sheet_names = test.sheetnames
    currencyStyle = NamedStyle(name='currencyStyle', number_format='0')
    for firm_name in test_sheet_names:
        for otchet_name in otchet_sheet_names:
            if firm_name == otchet_name:
                func_otchet(firm_name, otchet, test)
                otchet.save('2022 2 кв. - 2023г. 2 кв Продажи Общие — копия.xlsx')
                print(f'Обработанно: {firm_name}')
                break
        else:
            errors_file.write(firm_name)
            errors_file.write('  Лист не найден в отчёте')
            errors_file.write('\n')
